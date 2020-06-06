from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import time
import openpyxl
from haversine import haversine, Unit
os.chdir('c:\\Python\\colefitzpatrick_python\\FYA')

wb1 = openpyxl.load_workbook('fya_write_file.xlsx')       #file that comes from FARS
ws1 = wb1["Sheet1"]

wb2 = openpyxl.load_workbook('FYA_signals.xlsx') #intersection file
ws2 = wb2["FYA_Locations_FINAL"] #sheet name from intersection file

wb_write = openpyxl.load_workbook('linker_write.xlsx')
ws_write = wb_write["Sheet1"]

ws1numberrows = ws1.max_row
ws2numberrows = ws2.max_row
writerow = 2

for crashrow in range(2,ws1numberrows+1):
    distancelist = []
    lat1 = float(ws1.cell(row=crashrow, column=4).value) #lattitude
    long1 = float(ws1.cell(row=crashrow, column=5).value) #longitude
    crash_loc = (lat1, long1)
    for intersectionrow in range(2,ws2numberrows+1):
        lat2 = float(ws2.cell(row=intersectionrow, column=2).value) #intersection lat
        long2 = float(ws2.cell(row=intersectionrow, column=3).value) #intersection long
        intersection_loc = (lat2, long2)
        distance = haversine(crash_loc, intersection_loc, unit='ft')
        distancelist.append(distance)     
    index_min = min(range(len(distancelist)), key=distancelist.__getitem__)
    print(ws1.cell(row=crashrow, column=1).value)
    ws_write.cell(row=crashrow, column=1).value = ws1.cell(row=crashrow, column=1).value
    ws_write.cell(row=crashrow, column=2).value = ws1.cell(row=crashrow, column=2).value
    ws_write.cell(row=crashrow, column=3).value = index_min + 1
    ws_write.cell(row=crashrow, column=4).value = min(distancelist)

wb_write.save('linker_write.xlsx')
        
